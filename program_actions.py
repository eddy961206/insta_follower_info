
import os
import random
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime


def get_system_user_agent():
    """시스템의 현재 user-agent를 가져옵니다."""
    temp_driver = webdriver.Chrome()
    temp_driver.get("about:blank")
    user_agent = temp_driver.execute_script("return navigator.userAgent;")
    temp_driver.quit()
    return user_agent


def initialize_driver():
    """user-agent를 사용하여 크롬 드라이버 초기화"""
    user_agent = get_system_user_agent()
    options = Options()
    options.add_argument(f"user-agent={user_agent}")

    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(5)
    return driver


def initialize_driver_test():
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    driver = webdriver.Chrome(options=options)
    return driver


def save_to_excel(data):
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 시트 삭제

    # 각 계정별로 데이터 시트 생성
    for account, followers in data.items():
        ws = wb.create_sheet(title=account)
        # 헤더 추가
        headers = [f'{account} 계정', '팔로워 ID', '팔로워 수', '팔로잉 수', '계정 상태', '팔로워 링크']
        ws.append(headers)

        # 팔로워 데이터와 계정 이름, 순번 추가
        for index, follower in enumerate(followers, start=1):
            row_data = [index] + [follower.get(header, '') for header in headers[1:]]
            ws.append(row_data)

        # 컬럼 너비 조정
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    # 파일 이름 형식: 'YYMMDD_HHMM_인스타결과.xlsx'
    current_time = datetime.now().strftime('%y%m%d_%H%M')
    file_name = f"{current_time}_인스타결과.xlsx"
    wb.save(file_name)
    print(f"\n엑셀 파일 '{file_name}'에 데이터 저장 완료")



def find_element_with_retry(driver, by, value, delay=10):
    """요소가 나타날 때까지 찾기"""
    try:
        element = WebDriverWait(driver, delay).until(
            EC.element_to_be_clickable((by, value))
        )
        return element
    except TimeoutException:
        # return None
        raise NoSuchElementException(f"\n요소가 {delay}초 동안 나타나지 않았습니다: {value}")
    

# 에러 발생 시 스크린샷을 저장하는 함수
def capture_screenshot(driver, error_message):
    screenshots_dir = "에러 스크린샷"
    if not os.path.exists(screenshots_dir):
        os.makedirs(screenshots_dir)  # 스크린샷을 저장할 폴더가 없으면 생성
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')  # 현재 시간으로 타임스탬프 생성
    screenshot_file = os.path.join(screenshots_dir, f"error_{timestamp}.png")
    driver.save_screenshot(screenshot_file)  # 스크린샷 저장
    print(f"\n에러 스크린샷이 '{screenshot_file}' 파일에 저장되었습니다.")
    print(f"에러 메시지: {error_message}")



